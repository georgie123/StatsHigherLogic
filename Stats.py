
# Le script extrait les stats AMS à partir d'un export CSV des tables customer_ams et customer.

import os
from datetime import date, datetime
from tabulate import tabulate as tab
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
from matplotlib.collections import PatchCollection
from mpl_toolkits.basemap import Basemap
import numpy as np
from PIL import Image, ImageOps
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings
warnings.simplefilter(action='ignore', category=UserWarning)

today = date.today()

shp_simple_countries = r'C:/Users/Georges/OneDrive/_data/simple_countries/simple_countries'
shp_simple_areas = r'C:/Users/Georges/OneDrive/_data/simple_areas/simple_areas'
inputCountryConversion = r'C:/Users/Georges/OneDrive/_data/countries_conversion.xlsx'
inputSpecialtyConversion = r'C:/Users/Georges/OneDrive/_data/specialties_conversion.xlsx'

workDirectory = r'C:/Users/Georges/Downloads/'
outputExcelFile = workDirectory + str(today) + ' Stats AMS.xlsx'

dfAms = pd.read_csv(workDirectory + 'customer_ams.csv', sep=';', encoding='cp1252',
                            usecols=['id', 'customer_id', 'email', 'role'])

dfCustomer = pd.read_csv(workDirectory + 'customer.csv', sep=';', encoding='cp1252',
                            usecols=['id_customer', 'email', 'speciality', 'country'])
dfCustomer.rename(columns={'email': 'email_customer'}, inplace=True)

# print('\ncustomer_ams')
# print(tab(dfAms.head(10), headers='keys', tablefmt='psql', showindex=False))
# print(dfAms.shape[0])

# print('\nCustomer')
# print(tab(dfCustomer.head(10), headers='keys', tablefmt='psql', showindex=False))
# print(dfCustomer.shape[0])

# JOIN CUSTOMER_AMS AND CUSTOMER
df = pd.merge(dfAms, dfCustomer, left_on='customer_id', right_on='id_customer', how='left')[['id', 'role', 'email', 'speciality', 'country']]

print('\nAMS')
print(tab(df.head(10), headers='keys', tablefmt='psql', showindex=False))
print(df.shape[0])


# STATS SPECIALTY
df_Specialties = pd.DataFrame(df.groupby(['speciality'], dropna=False).size(), columns=['Total'])\
    .sort_values(['Total'], ascending=False).reset_index()
df_Specialties = df_Specialties.fillna('Unknow')

df_Specialties['Percent'] = (df_Specialties['Total'] / df_Specialties['Total'].sum()) * 100
df_Specialties['Percent'] = df_Specialties['Percent'].round(decimals=1)
df_Specialties.rename(columns={'speciality': 'Specialty'}, inplace=True)

print('\ndf_Specialties')
print(tab(df_Specialties.head(10), headers='keys', tablefmt='psql', showindex=False))
print(df_Specialties.shape[0])



# EXCEL FILE
writer = pd.ExcelWriter(outputExcelFile, engine='xlsxwriter')

df_Specialties.to_excel(writer, index=False, sheet_name='Specialties', header=True)

writer.save()




# EXCEL FILTERS
workbook = openpyxl.load_workbook(outputExcelFile)
sheetsLits = workbook.sheetnames

for sheet in sheetsLits:
    worksheet = workbook[sheet]
    FullRange = 'A1:' + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save(outputExcelFile)

# EXCEL COLORS
for sheet in sheetsLits:
    worksheet = workbook[sheet]
    for cell in workbook[sheet][1]:
        worksheet[cell.coordinate].fill = PatternFill(fgColor = 'FFC6C1C1', fill_type = 'solid')
        workbook.save(outputExcelFile)

# EXCEL COLUMN SIZE
for sheet in sheetsLits:
    for cell in workbook[sheet][1]:
        if get_column_letter(cell.column) == 'A':
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 30
        else:
            workbook[sheet].column_dimensions[get_column_letter(cell.column)].width = 10
        workbook.save(outputExcelFile)

workbook.save(outputExcelFile)

# EXCEL FREEZE TOP ROW
for sheet in sheetsLits:
    worksheet.freeze_panes = 'A2'



